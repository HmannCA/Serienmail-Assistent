import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from fastapi import UploadFile
from helpers import clean_for_json

UPLOAD_DIR = "user_uploads"

async def handle_excel_upload(excel_file: UploadFile, user_id: int) -> Dict[str, Any]:
    if not excel_file or not excel_file.filename:
        return {"error": "Keine Datei zum Hochladen ausgewählt."}

    file_info = os.path.splitext(excel_file.filename)
    extension = file_info[1].lower()
    
    if extension not in ['.xlsx', '.xls']:
        return {"error": "Ungültiges Dateiformat. Bitte .xlsx oder .xls hochladen."}

    new_filename = f"{user_id}_{os.urandom(8).hex()}{extension}"
    new_file_path = os.path.join(UPLOAD_DIR, new_filename)

    try:
        with open(new_file_path, "wb") as buffer:
            shutil.copyfileobj(excel_file.file, buffer)
        return {"file_path": new_file_path, "original_name": excel_file.filename}
    except Exception as e:
        return {"error": f"Fehler beim Speichern der hochgeladenen Datei: {e}"}

def read_excel_header(file_path: str) -> List[str]:
    header = []
    if not os.path.exists(file_path):
        return []
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        for cell in sheet[1]:
            if cell.value is not None and str(cell.value).strip() != '':
                header.append(str(clean_for_json(cell.value)).strip())
        workbook.close()
    except Exception as e:
        raise Exception(f"Kritischer Fehler beim Lesen der Kopfzeile der Excel-Datei: {e}")
    return header

def read_all_excel_data(file_path: str) -> List[Dict[str, Any]]:
    all_data = []
    if not os.path.exists(file_path):
        return []

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        header_raw = [str(clean_for_json(cell.value)).strip() for cell in sheet[1] if cell.value is not None and str(cell.value).strip() != '']
        header_index_map = {name: idx + 1 for idx, name in enumerate(header_raw)}

        for row_index in range(2, sheet.max_row + 1):
            row_data = {}
            for col_name, col_idx in header_index_map.items():
                cell_obj = sheet[f"{get_column_letter(col_idx)}{row_index}"]
                row_data[col_name] = clean_for_json(cell_obj.value)
            
            if any(value is not None and str(value).strip() != '' for value in row_data.values()):
                all_data.append(row_data)

        workbook.close()
    except Exception as e:
        raise Exception(f"Kritischer Fehler beim Lesen der Daten: {e}")

    return all_data

# === KORRIGIERTE FILTER-FUNKTION ===
def filter_excel_data(file_path: str, column_name: str, filter_value: str) -> List[Dict[str, Any]]:
    filtered_data = []
    if not os.path.exists(file_path):
        return []

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        header_raw = [str(clean_for_json(cell.value)).strip() for cell in sheet[1] if cell.value is not None and str(cell.value).strip() != '']
        header_index_map = {name: idx + 1 for idx, name in enumerate(header_raw)}
        
        column_index_to_filter = header_index_map.get(column_name)
        if column_index_to_filter is None:
            raise ValueError(f"Filter-Spalte '{column_name}' nicht in Excel-Header gefunden.")

        column_letter_to_filter = get_column_letter(column_index_to_filter)

        for row_index in range(2, sheet.max_row + 1):
            cell_value_on_row = sheet[f"{column_letter_to_filter}{row_index}"].value

            # Hier ist die entscheidende Korrektur:
            # 1. Prüfen, ob der Wert in der Zelle überhaupt vorhanden ist (nicht None ist).
            # 2. Beide Werte (aus der Zelle und vom Nutzer) in Text umwandeln, Leerzeichen entfernen und in Kleinbuchstaben umwandeln für einen robusten Vergleich.
            if cell_value_on_row is not None and str(cell_value_on_row).strip().lower() == filter_value.strip().lower():
                row_data = {}
                for col_name, col_idx in header_index_map.items():
                    cell_obj = sheet[f"{get_column_letter(col_idx)}{row_index}"]
                    row_data[col_name] = clean_for_json(cell_obj.value)
                
                if any(value is not None and str(value).strip() != '' for value in row_data.values()):
                    filtered_data.append(row_data)
        workbook.close()
    except Exception as e:
        raise Exception(f"Kritischer Fehler beim Filtern der Daten: {e}")

    return filtered_data